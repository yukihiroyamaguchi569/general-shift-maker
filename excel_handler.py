"""
エクセルファイルの読み込み・書き出し（当直表フォーマット）

フォーマット:
  行1: ヘッダー（A1: 日直回数上限, B1: 当直回数上限, C1: 職員番号, D1以降: 日付）
  行2以降: 職員データ（A列: 日直上限, B列: 当直上限, C列: 職員番号, D列以降: シフト）

※ 休診日はUIのプルダウンで設定（Excel内に休診日行なし）
"""

import csv
from io import BytesIO, StringIO
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================
# シフト種別ごとの色設定
# =========================================================

SHIFT_COLORS = {
    "当直": {"bg": "1e3a5f", "fg": "FFFFFF"},  # 濃紺背景・白文字
    "日直": {"bg": "ffcdd2", "fg": "333333"},  # 薄赤背景・黒文字
}

# 固定値（☓, 出, 休 など）はグレー
FIXED_COLOR = {"bg": "e0e0e0", "fg": "666666"}


def read_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    エクセルファイルを読み込んでJSON用の辞書を返す

    フォーマット:
      新: col1=日直上限, col2=当直上限, col3=合計上限, col4=職員番号, col5+=日付
      旧: col1=日直上限, col2=当直上限, col3=職員番号, col4+=日付（後方互換）

    Returns:
        {
            "staff_ids": ["01", "02", ...],
            "day_limits": [2, 0, ...],
            "night_limits": [4, 6, ...],
            "total_limits": [8, 6, ...],
            "dates": [1, 2, ..., 31],
            "schedule": [["", "☓", ...], ...],
            "closed_days": [],
        }
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # フォーマット検出: 行2のcol3が数値なら新フォーマット
    new_format = False
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, 3).value
        if val is not None and str(val).strip():
            try:
                int(str(val).strip())
                new_format = True
            except ValueError:
                new_format = False
            break

    name_col = 4 if new_format else 3
    date_col_start = 5 if new_format else 4

    dates: List[Any] = []
    for col in range(date_col_start, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is None:
            break
        try:
            day_num = int(val)
            if 1 <= day_num <= 31:
                dates.append(day_num)
            else:
                break
        except (ValueError, TypeError):
            break

    num_date_cols = len(dates)

    # 行2以降: 職員データ
    staff_ids: List[str] = []
    day_limits: List[int] = []
    night_limits: List[int] = []
    total_limits: List[int] = []
    schedule: List[List[str]] = []

    for row_idx in range(2, ws.max_row + 1):
        staff_id = ws.cell(row_idx, name_col).value
        if staff_id is None or str(staff_id).strip() == "":
            continue

        day_lim_val = ws.cell(row_idx, 1).value
        try:
            day_lim = int(day_lim_val) if day_lim_val is not None else 0
        except (ValueError, TypeError):
            day_lim = 0

        night_lim_val = ws.cell(row_idx, 2).value
        try:
            night_lim = int(night_lim_val) if night_lim_val is not None else 0
        except (ValueError, TypeError):
            night_lim = 0

        if new_format:
            total_lim_val = ws.cell(row_idx, 3).value
            try:
                total_lim = int(total_lim_val) if total_lim_val is not None else 99
            except (ValueError, TypeError):
                total_lim = 99
        else:
            total_lim = 99

        staff_ids.append(str(staff_id).strip())
        day_limits.append(day_lim)
        night_limits.append(night_lim)
        total_limits.append(total_lim)

        row_data: List[str] = []
        for col in range(date_col_start, date_col_start + num_date_cols):
            val = ws.cell(row_idx, col).value
            if val is None:
                row_data.append("")
            else:
                row_data.append(str(val).strip())
        schedule.append(row_data)

    return {
        "staff_ids": staff_ids,
        "day_limits": day_limits,
        "night_limits": night_limits,
        "total_limits": total_limits,
        "dates": dates,
        "schedule": schedule,
        "closed_days": [],
    }


def _detect_csv_format(rows: list) -> bool:
    """3列目が数値なら新フォーマット（合計上限列あり）と判定"""
    for row in rows[1:]:
        if len(row) >= 3 and row[2].strip():
            try:
                int(row[2].strip())
                return True
            except ValueError:
                return False
    return False


def read_csv(file_bytes: bytes) -> Dict[str, Any]:
    """
    CSVファイルを読み込んでJSON用の辞書を返す

    CSVフォーマット:
      新: 日直回数上限, 当直回数上限, 合計上限, 職員番号, 日付...
      旧: 日直回数上限, 当直回数上限, 職員番号, 日付...（後方互換）

    Returns:
        read_excel() と同じ辞書構造（total_limits を含む）
    """
    # utf-8 → cp932 の順でデコードを試みる
    text: str = ""
    for encoding in ("utf-8-sig", "utf-8", "cp932"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    reader = csv.reader(StringIO(text))
    rows = list(reader)

    if not rows:
        return {"staff_ids": [], "day_limits": [], "night_limits": [], "total_limits": [], "dates": [], "schedule": [], "closed_days": []}

    # フォーマット検出
    new_format = _detect_csv_format(rows)
    name_col = 3 if new_format else 2
    date_col_start = 4 if new_format else 3

    # 日付列を読む
    header = rows[0]
    dates: List[Any] = []
    for col in range(date_col_start, len(header)):
        val = header[col].strip()
        if not val:
            break
        try:
            day_num = int(val)
            if 1 <= day_num <= 31:
                dates.append(day_num)
            else:
                break
        except (ValueError, TypeError):
            break

    num_date_cols = len(dates)

    # 行2以降: 職員データ
    staff_ids: List[str] = []
    day_limits: List[int] = []
    night_limits: List[int] = []
    total_limits: List[int] = []
    schedule: List[List[str]] = []

    for row in rows[1:]:
        if len(row) <= name_col:
            continue
        staff_id = row[name_col].strip()
        if not staff_id:
            continue

        try:
            day_lim = int(row[0].strip()) if row[0].strip() else 0
        except (ValueError, TypeError):
            day_lim = 0

        try:
            night_lim = int(row[1].strip()) if row[1].strip() else 0
        except (ValueError, TypeError):
            night_lim = 0

        if new_format:
            try:
                total_lim = int(row[2].strip()) if row[2].strip() else 99
            except (ValueError, TypeError):
                total_lim = 99
        else:
            total_lim = 99

        staff_ids.append(staff_id)
        day_limits.append(day_lim)
        night_limits.append(night_lim)
        total_limits.append(total_lim)

        row_data: List[str] = []
        for col in range(date_col_start, date_col_start + num_date_cols):
            val = row[col].strip() if col < len(row) else ""
            row_data.append(val)
        schedule.append(row_data)

    return {
        "staff_ids": staff_ids,
        "day_limits": day_limits,
        "night_limits": night_limits,
        "total_limits": total_limits,
        "dates": dates,
        "schedule": schedule,
        "closed_days": [],
    }


def write_excel(
    staff_ids: List[str],
    day_limits: List[int],
    night_limits: List[int],
    total_limits: List[int],
    dates: List[Any],
    schedule: List[List[str]],
    closed_days: List[int],
    sheet_title: str = "当直表",
) -> bytes:
    """
    当直表データからエクセルファイルを生成してバイト列で返す
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    # 行1: ヘッダー
    headers = ["日直回数上限", "当直回数上限", "合計上限", "職員番号"]
    for i, h in enumerate(headers):
        cell = ws.cell(1, i + 1, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    for i, date_val in enumerate(dates):
        col = i + 5
        cell = ws.cell(1, col, date_val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    # 行2以降: 職員データ
    for row_idx, staff_id in enumerate(staff_ids):
        excel_row = row_idx + 2

        cell = ws.cell(excel_row, 1, day_limits[row_idx] if row_idx < len(day_limits) else 0)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align

        cell = ws.cell(excel_row, 2, night_limits[row_idx] if row_idx < len(night_limits) else 0)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align

        cell = ws.cell(excel_row, 3, total_limits[row_idx] if row_idx < len(total_limits) else 99)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align

        cell = ws.cell(excel_row, 4, staff_id)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align

        for col_idx, shift in enumerate(schedule[row_idx]):
            col = col_idx + 5
            cell = ws.cell(excel_row, col, shift)
            cell.border = thin_border
            cell.alignment = center_align

            if shift in SHIFT_COLORS:
                colors = SHIFT_COLORS[shift]
                cell.fill = PatternFill(
                    start_color=colors["bg"],
                    end_color=colors["bg"],
                    fill_type="solid",
                )
                cell.font = Font(size=10, color=colors["fg"], bold=(shift == "当直"))
            elif shift.strip() != "":
                cell.fill = PatternFill(
                    start_color=FIXED_COLOR["bg"],
                    end_color=FIXED_COLOR["bg"],
                    fill_type="solid",
                )
                cell.font = Font(size=10, color=FIXED_COLOR["fg"])
            else:
                cell.font = cell_font

    # フリーズペイン（1行目とD列を固定）
    ws.freeze_panes = "E2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
